# this is a script file for name tags
# author: Hong Yoon Kim
# moderation and upkeep: Caressa Chan
"""
letter: 612, 792
individual name tag size: 288 x 216
grid: x, (18,306,594,612)
	y, (72,288,504,720)
"""
from reportlab.platypus import PageBreak, Image
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import landscape, letter, A4
from PIL import Image
from xlrd import open_workbook
import time
from operator import itemgetter
from openpyxl.styles import Font       #Style, Font
from openpyxl.cell import Cell
import xlrd
from reportlab.lib.colors import PCMYKColor, PCMYKColorSep, Color, black
from reportlab.lib.units import inch

# THIS IS THE DREAM
# To have nothing specific in the code
# So people can just drag and select and it'll be cool

# TODO Step 1: Register all fonts that you will be using
pdfmetrics.registerFont(TTFont('Raleway Regular','Raleway-Regular.ttf'))
pdfmetrics.registerFont(TTFont('Raleway Bold','Raleway-Bold.ttf'))

# TODO Step 2: Change your input filename, choose which tab (sheet) you'll
# 			   be reading from, and name the important columns
def ingestXLS(xlsFile):
	filename = 'nametags.xls'
	book = open_workbook(filename, formatting_info=True)
	sheet = book.sheet_by_name('Sheet1')
	font = book.font_list
	nrows = sheet.nrows
	print 'number of row is', nrows
	ncols = sheet.ncols
	print 'number of col is', ncols
	data = []
	# skip the first row cuz it's just labels of what is it
	for row in range(1,nrows):                       # autoincrement at end
	   personName = sheet.cell(row, 0).value
	   lifeGroup = sheet.cell(row, 1).value
	   housing = sheet.cell(row, 2).value
	   keyHolder = sheet.cell(row, 3).value
	   smallGroup = sheet.cell(row, 6).value

	   # finding keyholder
	   #cell_xf = book.xf_list[sheet.cell_xf_index(row,0)]
	   #if font[cell_xf.font_index].bold: housing = housing + '.'
	   if keyHolder == 'x': housing = housing + '*'

	   datum = (personName, lifeGroup, housing, smallGroup)
	   data.append(datum)

	# sort by location (1) then name (0)
	#data = sorted(data, key=itemgetter(1,0))
	return data

# TODO Step 3: Choose the font you'll be using for the name
def getNameFontSize(canvas, name):
	#generate strings of names that will be on the name tag.
	fieldWidth = 252
	nameFontSize = 45
	nameFont='Raleway Bold'
	threshold = fieldWidth*.90
	sw = canvas.stringWidth(name,nameFont,nameFontSize)

	while sw > threshold:
		nameFontSize = nameFontSize - 0.1
		sw = canvas.stringWidth(name,nameFont,nameFontSize)

		# if nameFontSize < 30:
		# 	nameSplit = name.split()
		# 	lastName = nameSplit[-1][0] + '.'
		# 	print('last name is replaced to the first letter ' , name )
		# 	nameSplit.pop()

		# 	if len(nameSplit) > 1:
		# 		firstName = ' '.join(nameSplit)
		# 	else:
		# 		firstName = nameSplit[0]

		# 	name = ' '.join([firstName, lastName])
		# 	nameFontSize = 40
		# 	sw = canvas.stringWidth(name,nameFont,nameFontSize)

		# 	while sw > threshold:
		# 		nameFontSize = nameFontSize - 0.1
		# 		sw = canvas.stringWidth(name,nameFont,nameFontSize)

	return nameFontSize

# TODO Step 4: Change background image name, choose fonts for all other labels
#			   Also manually change locations (x and ycoords)of other labels 
def genSingleNameTag(canvas, datum, xcoord, ycoord):
	# generate a single name tag
	# dimensions: 252 x 160

	# print background image
	#im = Image.open('newbackground.jpg')
	myimg = 'final_devoted_bg.jpg'
	canvas.drawImage(myimg, xcoord, ycoord, width=252, height=160)

	# unpack datum
	# print datum
	personName, lifeGroup, housing, smallGroup = datum   # might include keyholder later

	# print personName
	# modify name / fontSize if need
	nameFontSize = getNameFontSize(canvas,personName)
	canvas.setFont('Raleway Bold',nameFontSize)
	canvas.setFillColorRGB(0, 0, 0)
	canvas.drawCentredString(xcoord+126, ycoord + 85 -nameFontSize/2,personName)   #ycoord was 70

	# print housing
	canvas.setFont('Raleway Bold', 15)
	canvas.setFillColorRGB(0, 0, 0)
	canvas.drawString(xcoord+5,ycoord+5,housing)   #240,12

	# print lifeGroup (this will be where they're from)
 	canvas.setFont('Raleway Regular', 20) #24
	canvas.setFillColorRGB(0, 0, 0)
	canvas.drawCentredString(xcoord+126,ycoord+40,lifeGroup) #12, 12

	# print smallGroup
	# check if they're actually in a small group?
	canvas.setFont('Raleway Bold', 15)
	canvas.setFillColorRGB(0, 0, 0)
	canvas.drawString(xcoord+180, ycoord+5,smallGroup)

def genSinglePageNameTag(canvas, data):
	print 'new page!'
	xcoords = [54, 306]
   	ycoords = [76, 236, 396, 556]

   	for x in xcoords:
   		for y in ycoords:
   			if data:
   				datum = data.pop(0)
   				genSingleNameTag(canvas,datum,x,y)
			else:
				break
 	canvas.setStrokeColor(black)
 	canvas.grid([54, 306, 558], [76, 236, 396, 556, 716])

def genPdfNameTag(canvas, data):
	while data:
		if len(data) > 8:
			# select first 8 elements
			genSinglePageNameTag(canvas, data[:8])

			# truncate data
			data = data[8:len(data)]

			# move to the next page
			canvas.showPage()
		else:
			genSinglePageNameTag(canvas, data)


def printPDF(excelFile, pdfName = "nameTag.pdf"):
	data = ingestXLS(excelFile)
	c = canvas.Canvas(pdfName,pagesize = letter)
	genPdfNameTag(c,data)
 	c.setStrokeColor(black)
 	c.grid([54, 306, 558], [76, 236, 396, 556, 716])
	c.save()

start = time.time()
printPDF('nameTag.xls')
end = time.time()

print 'it lasts ',end-start, ' seconds'

# def generateNameTag(xlsFile, pdfName = "meSoLazy.pdf"):
# 	data = ingestXLS(xlsFile)

# 	for datum in data:
# 		c = canvas.Canvas(pdfName, pagesize = letter)
# genPdfNameTag(c,names, housings)
# c.save()
# print('it is saved')
# return
